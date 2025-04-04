import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from typing import List, Optional
from pydantic import ValidationError
from app.config import get_config
from app.utils import rp
import numpy as np
from datetime import datetime, timedelta
import random
from math import ceil
class RuleService:
    def __init__(self):
        self.conf = get_config()
        self.oil_mapping = self.conf.get("BUSINESS.RESTAURANT.收油关系映射", default={})

    """
    生成收油表步骤，字段名还需要修改
    """

    ## 读取配置文件收油关系映射
    def oil_determine_collection_amount(self,restaurant_type: str, oil_mapping: dict) -> int:
        for key, value in oil_mapping.items():
            if ',' in str(value):
                allocate_value =  [int(item.strip()) for item in str(value).split(',')]
            else:
                allocate_value = [int(value)]
            if any(type_keyword in restaurant_type for type_keyword in key.split('/')):
                return np.random.choice(allocate_value)
        return np.random.choice([1, 2])  # 默认值

    """
    收油表
    传入餐厅信息，包括required_columns中的列,并根据餐厅类型增加收油数的列，
    最后返回排序后的餐厅信息
    """
    def oil_restaurant_sort(self,df: pd.DataFrame) -> pd.DataFrame:

        # 确保需要的列存在
        required_columns = ['Chinese name', 'English name', 'Chinese Address', 'English Address',
                            'Coordinates', 'Contact person(EN)', 'Telephone number', 'Distance (km)',
                            '镇/街道', '区域', '餐厅类型']
        
        if not all(column in df.columns for column in required_columns):
            raise ValueError("DataFrame缺少必要的列")

        # 增加一列随机数作为桶数
        df['桶数'] = np.random.rand(df.shape[0])
        
        # 根据区域，镇/街道、桶数进行排序
        df_sorted = df.sort_values(by=['区域', '镇/街道', '桶数'])

        
        # 应用函数到DataFrame以生成收油数列
        df_sorted['收油数'] = df_sorted['餐厅类型'].apply(lambda x :self.oil_determine_collection_amount(x,self.oil_mapping))
        
        return df_sorted
    
    """
    分配车辆号码
    传入餐厅信息和车辆信息，根据收油数分配车辆号码，并将结果与原DataFrame合并
    """
    def oil_assign_vehicle_numbers(self,df_restaurants: pd.DataFrame, df_vehicles: pd.DataFrame,total_barrels: int) -> pd.DataFrame:
        """
        根据收油数分配车辆号码，并将结果与原DataFrame合并
        
        :param df_restaurants: 包含'镇/街道', '区域', '餐厅类型', '收油数'的DataFrame
        :param df_vehicles: 包含'车牌号'的DataFrame
        :param total_barrels: 总桶数限制
        :return: 处理后的DataFrame
        """
        # 初始化变量
        vehicle_numbers = df_vehicles['车牌号'].sample(frac=1, replace=False).tolist()  # 乱序车牌号
        result_rows = []
        current_vehicle_index = 0
        total_accumulated = 0  # 所有车辆的累计收油数
        should_break = False  # 控制外层循环的标志
        
        # 按区域分组
        grouped = df_restaurants.groupby('区域')
        
        for area, group in grouped:
            
            accumulated_sum = 0
            temp_group = []
            
            for index, row in group.iterrows():
                temp_group.append(row)
                accumulated_sum += row['收油数']
                
                # 如果累计值达到在35-44之间，则分配车牌号并重置累计值
                if accumulated_sum>=35:
                    # 检查添加这组数据是否会超过总桶数限制
                    if total_accumulated > total_barrels:
                        should_break = True  # 设置跳出标志
                        break
                        
                    # 分配车牌号
                    for temp_row in temp_group:
                        temp_row = temp_row.copy()  # 防止修改原DataFrame
                        temp_row['车牌号'] = vehicle_numbers[current_vehicle_index]
                        temp_row['累计收油数'] = accumulated_sum
                        result_rows.append(temp_row)
                    
                    total_accumulated += accumulated_sum
                    print(f"当前累计桶数: {total_accumulated}, 目标桶数: {total_barrels}")  # 调试信息
                
                    current_vehicle_index = (current_vehicle_index + 1) % len(vehicle_numbers)
                    accumulated_sum = 0
                    temp_group = []
            
            # 如果外层循环需要跳出，则不处理剩余数据
            if should_break:
                break
            

     
        # 创建结果DataFrame
        result_df = pd.DataFrame(result_rows)        
        # 打印统计信息
        print("\n统计信息:")
        print(f"总分配桶数: {total_accumulated}")
        print(f"总桶数限制: {total_barrels}")
        print(f"已分配车辆数: {len(set(result_df['车牌号']))}")
        print(f"已处理餐厅数: {len(result_rows)}")
        print(f"总餐厅数: {len(df_restaurants)}")
        
        return result_df

    """
    写入Excel文件，并合并分配的车牌号和对应的累加收油数单元格
    """
    def oil_write_to_excel_with_merge_cells(self,df: pd.DataFrame, output_path: str):
        """
        将DataFrame写入Excel文件，并合并分配的车牌号和对应的累加收油数单元格
        
        :param df: 要写入Excel的DataFrame
        :param output_path: 输出Excel文件路径
        """
        wb = Workbook()
        ws = wb.active
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 合并单元格
        start_row = 2  # 假设第一行是标题
        while start_row <= ws.max_row:
            end_row = start_row
            while end_row < ws.max_row and ws[f'E{end_row}'].value == ws[f'E{end_row + 1}'].value:
                end_row += 1
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)  # 车牌号列
                ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)  # 累计收油数列
                ws[f'F{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'G{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            start_row = end_row + 1
        
        wb.save(output_path)

    """
    生成平衡表-五月表步骤
    输入收油表，包含'区域', '车牌号', '累计收油数'字段；输入
    步骤1：读取dataframe中的'区域', '车牌号', '累计收油数'字段作为新的dataframe的字段，并去重
    步骤2：新建一列榜单净重，公式为累计收油数*0.18-RANDBETWEEN(1,5)/100
    步骤3：新建几列固定值的列
    步骤4：计算车数car_number_of_day并新增交付日期列
    步骤5：计算加工量和毛油库存
    步骤6：计算产出重量
    步骤7：计算辅助列
    """

    def process_balance_dataframe(self,df: pd.DataFrame, n: int,current_date: str) -> pd.DataFrame:
        """
        根据给定的步骤处理输入的DataFrame。
        
        :param df: 输入的DataFrame，包含'区域', '车牌号', '累计收油数'字段
        :param n: 多少天运完
        :return: 处理后的DataFrame
        """
        # 步骤1：读取dataframe中的'区域', '车牌号', '累计收油数'字段作为新的dataframe的字段，并去重
        new_df = df[['区域', '车牌号', '累计收油数']].drop_duplicates()
        new_df['收集城市'] = new_df['区域']
        # 步骤2：新建一列榜单净重，公式为累计收油数*0.18-RANDBETWEEN(1,5)/100
        new_df['榜单净重'] = new_df['累计收油数'].apply(lambda x: x * 0.18 - random.randint(1, 5) / 100)

        # 步骤3：新建几列固定值的列
        current_year_month = datetime.strptime(current_date, '%Y-%m-%d')
        new_df['货物类型'] = '餐厨废油'
        new_df['运输方式'] = '大卡车'
        new_df['流水号'] = [f"{current_year_month}{str(i+1).zfill(3)}" for i in range(len(new_df))]
        new_df['榜单编号'] = 'B' + new_df['流水号']

        # 步骤4：计算车数car_number_of_day并新增交付日期列
        car_number_of_day = ceil(len(new_df) // n) ## 计算每天大概需要多少辆车
        dates_in_month = pd.date_range(start=datetime(datetime.now().year, datetime.now().month, 1), 
                                    end=(datetime(datetime.now().year, datetime.now().month, 1) + pd.offsets.MonthEnd(0)))
        delivery_dates = []
        day_index = 0
        
        while day_index < len(dates_in_month):
            delta = car_number_of_day + random.choice([-1, 0, 1])
            if delta <= 0:
                delta = 1  # 确保至少有一辆车
            for _ in range(min(delta, len(new_df) - len(delivery_dates))):
                delivery_dates.append(dates_in_month[day_index].date())
            day_index += 1
        
        # 如果生成的交付日期少于新数据框的行数，则用最后一天填充剩余部分
        if len(delivery_dates) < len(new_df):
            last_date = delivery_dates[-1] if delivery_dates else dates_in_month[-1].date()
            delivery_dates.extend([last_date] * (len(new_df) - len(delivery_dates)))
        
        new_df['交付日期'] = delivery_dates[:len(new_df)]
        
        return new_df
    """
    总表：生成毛油库存 期末库存、辅助列、转化系数、产出重量、售出数量、加工量
    传入平衡表_5月表
    售出数量需要修改，规则没确定
    步骤1：复制日期、车牌号、榜单净重、榜单编号、收集城市;
    步骤2：新增一列加工量，如果当前日期为相同日期的最后一行，则加工量等于每个日期的榜单净重和，否则等于0 ；
        新增一列毛油库存，公式为当前行的榜单净重+上一行的毛油库存-当前行的加工量；
        新增一列辅助列，值为当前日期如果等于下一行的日期，则为空值，否则为1；
    新增1列转化系数，值为=RANDBETWEEN(900,930)/1；
    新增1列产出重量，值为round(加工量*转化系数/100,2);
    新增1列售出数量，值为0
    """
    def process_dataframe_with_new_columns(self, df: pd.DataFrame, total_df: pd.DataFrame = None) -> pd.DataFrame:
        """
        处理DataFrame并与总表合并
        
        :param df: 当月平衡表，输入的DataFrame，至少包含'日期', '车牌号', '榜单净重', '榜单编号', '收集城市'字段
        :param total_df: 总表DataFrame，如果为None则创建新的DataFrame
        :return: 处理后的DataFrame
        """
        # 步骤1：新建一个dataframe,从dataframe复制日期、车牌号、榜单净重、榜单编号、收集城市
        new_df = df[['日期', '车牌号', '榜单净重', '榜单编号', '收集城市']].copy()
        
        # 初始化新增的列
        new_df['加工量'] = 0.0
        new_df['毛油库存'] = 0.0
        new_df['辅助列'] = None
        new_df['转化系数'] = [np.random.randint(900, 931) for _ in range(len(new_df))]
        new_df['产出重量'] = round(new_df['加工量'] * new_df['转化系数'] / 100, 2)
        new_df['售出数量'] = 0
        new_df['期末库存'] = 0.0
        
        # 步骤2：计算加工量和毛油库存
        for date in new_df['日期'].unique():
            mask = new_df['日期'] == date
            if mask.sum() > 0:  # 确保有数据
                total_weight = new_df.loc[mask, '榜单净重'].sum()
                last_index = new_df[mask].index[-1]
                new_df.at[last_index, '加工量'] = total_weight
                
                # 更新毛油库存
                running_total = 0.0
                for idx in new_df[mask].index:
                    current_weight = new_df.at[idx, '榜单净重']
                    previous_inventory = running_total if idx != new_df[mask].index[0] else 0
                    processing_amount = new_df.at[idx, '加工量']
                    new_df.at[idx, '毛油库存'] = current_weight + previous_inventory - processing_amount
                    running_total = new_df.at[idx, '毛油库存']
                    
        # 计算产出重量
        new_df['产出重量'] = round(new_df['加工量'] * new_df['转化系数'] / 100, 2)
        
        # 计算辅助列
        new_df['辅助列'] = new_df['日期'].ne(new_df['日期'].shift(-1)).astype(int)
        new_df['辅助列'] = new_df['辅助列'].replace({1: 1, 0: None})
        
        # 计算期末库存
        previous_end_stock = 0.0 #第一行的期末库存前一行默认0
        for index, row in new_df.iterrows():
            current_output = row['产出重量']
            current_sale = row['售出数量']
            new_df.at[index, '期末库存'] = current_output + previous_end_stock - current_sale
            previous_end_stock = new_df.at[index, '期末库存']

        # 如果提供了总表，则合并数据
        if total_df is not None:
            # 使用concat合并两个DataFrame
            result_df = pd.concat([total_df, new_df], ignore_index=True)
            return result_df
        
        return new_df
    """
    收货确认书，传入收油表、销售车牌信息、收油重量、收货确认天数
    """
    def generate_df_check(self, oil_weight: float, days: int, df_oil: pd.DataFrame, df_car: pd.DataFrame,currrnt_date:str) -> pd.DataFrame:
        """
        生成检查数据表
        
        :param oil_weight: 收油重量（吨）
        :param days: 天数
        :param df_oil: 收油表DataFrame
        :param df_car: 销售车牌表DataFrame
        :param current_date: 当前月
        :return: 生成的检查数据表DataFrame
        """
        def random_weight():
            """生成随机重量（吨）"""
            return np.random.randint(3050, 3496) / 100
        
        def get_difference_value():
            """查表获取差值"""
            lookup_keys = [0,3,6,10,15,30,60,90,150,200,300,350,480,550,700,800,850,900,940,970,990,995,1001]
            lookup_values = [-15,-14,-13,-12,-11,-7,-6,-5,-4,-3,-2,-1,0,1,2,3,4,5,6,7,11,12]
            random_num = np.random.randint(1, 1001)
            
            # 找到对应区间
            for i, key in enumerate(lookup_keys):
                if random_num <= key:
                    return lookup_values[i] / 100
            return 0

        # 步骤1：创建df_check基本结构
        df_check = pd.DataFrame(columns=[
            '提货日期', '名称', '车牌号', '重量', '司机', '磅单号',
            '毛重', '皮重', '净重', '卸货重量', '差值'
        ])
    
        # 步骤2：确定行数和重量
        total_weight = 0
        weights = []
        while True:
            weight = random_weight()
            weights.append(weight)
            total_weight = sum(weights)
                # 检查是否在目标重量的±5%范围内
            if total_weight >= oil_weight:
                lower_bound = oil_weight * 0.95
                upper_bound = oil_weight * 1.05
                if lower_bound <= total_weight <= upper_bound:
                    break
                else:
                    # 重新开始
                    weights = []
                    total_weight = 0
        rows_count = len(weights)
    
        # 步骤3：分配日期
        start_date = pd.to_datetime(df_oil['提货日期'].min()) + timedelta(days=1)
        base_cars_per_day = rows_count // days  # 基本每日车次
        remaining_total_cars = rows_count       # 剩余需要分配的总车次
        dates = []

        for day in range(days):
            current_date = start_date + timedelta(days=day)
            
            if day == days - 1:  # 最后一天
                # 将剩余所有车次分配到最后一天
                day_cars = remaining_total_cars
            else:
                # 随机生成当天车次 (基础车次 ± 1)
                day_cars = base_cars_per_day + np.random.choice([-1, 0, 1])
                # 确保不会分配过多或过少
                if remaining_total_cars - day_cars < (days - day - 1):  # 确保后面的天数至少每天能分配1辆车
                    day_cars = remaining_total_cars - (days - day - 1)
                elif remaining_total_cars - day_cars > (days - day - 1) * (base_cars_per_day + 1):  # 确保后面的天数不会超出最大可能车次
                    day_cars = remaining_total_cars - (days - day - 1) * (base_cars_per_day + 1)
            
            # 确保当天车次不会小于0
            day_cars = max(1, min(day_cars, remaining_total_cars))
            remaining_total_cars -= day_cars
            
            # 添加当天的日期
            dates.extend([current_date] * day_cars)
        
        # 步骤5：分配车辆信息
        # 随机打乱车辆信息
        df_car_shuffled = df_car.sample(frac=1).reset_index(drop=True)
        car_info = df_car_shuffled.iloc[:(rows_count % len(df_car_shuffled) + 1)].copy()
        while len(car_info) < rows_count:
            car_info = pd.concat([car_info, df_car_shuffled])
        car_info = car_info.iloc[:rows_count]
        
        # 创建最终DataFrame
        df_check = pd.DataFrame({
            '提货日期': dates,
            '重量': weights,
            '车牌号': car_info['车牌号'].values,
            '司机': car_info['司机'].values,
            '皮重': car_info['皮重'].values + np.random.randint(1, 14, size=rows_count) * 10
        })
        
        # 步骤6：计算其他列
        df_check['净重'] = (df_check['重量'] * 1000).astype(int)
        df_check['毛重'] = df_check['皮重'] + df_check['净重']
        df_check['差值'] = [get_difference_value() for _ in range(len(df_check))]
        
        # 步骤7：计算卸货重量
        df_check['卸货重量'] = df_check['重量'] + df_check['差值']
        
        # 生成磅单号（示例：使用日期和序号组合）
        current_date = datetime.strptime(current_date, '%Y-%m-%d').strftime('%Y%m')
        df_check['磅单号'] = df_check.apply(
            lambda row: f"BD{current_date.strftime('%Y%m')}{str(row.name+1).zfill(3)}", 
            axis=1
        )
        
        # 添加名称列（如果需要可以根据实际需求修改）
        df_check['名称'] = "工业级混合油"
        
        # 调整列顺序
        df_check = df_check[[
            '提货日期', '名称', '车牌号', '重量', '司机', '磅单号',
            '毛重', '皮重', '净重', '卸货重量', '差值'
        ]]
        
        return df_check
    
    """
    复制收货确认书的"数据透视表"的每日重量一列至物料平衡表-总表，对齐日期
    传入收货确认书和平衡表-总表
    """
    def process_check_to_sum(df_generate_check: pd.DataFrame, df_generate_sum: pd.DataFrame) -> pd.DataFrame:
        """
        处理两个DataFrame并生成一个新的DataFrame。
        
        :param df_generate_check: 包含提货日期和重量的DataFrame,收货确认书
        :param df_generate_sum: 包含供应日期和售出数量的DataFrame，物料平衡表-总表
        :return: 处理后的DataFrame
        """
        # 步骤1：对df_generate_check表根据提货日期对重量进行求和汇总得到df_sum
        df_sum = df_generate_check.groupby('提货日期')['重量'].sum().reset_index()
        df_sum.rename(columns={'重量': '汇总重量'}, inplace=True)
        
        # 步骤2：关联df_generate_sum和df_sum，根据df_generate_sum的供应日期和df_sum的提货日期，
        # 将df_generate_sum中相同日期的最后一行的售出数量赋值为df_sum对应日期的汇总值
        df_merged = pd.merge(df_generate_sum, df_sum, left_on='供应日期', right_on='提货日期', how='left')
        
        # 对于每个日期，找到最后一行并将售出数量设置为汇总重量
        for date in df_merged['供应日期'].unique():
            mask = df_merged['供应日期'] == date
            if mask.sum() > 0:  # 确保有数据
                last_index = df_merged[mask].index[-1]  # 获取相同日期中的最后一行索引
                df_merged.at[last_index, '售出数量'] = df_merged.loc[last_index, '汇总重量']
        
        # 删除不必要的列
        df_final = df_merged.drop(columns=['提货日期', '汇总重量'])
        
        return df_final
    
    """
    复制平衡表-5月表的"流水号 车牌号 交付时间"到收油表的"流水号 车牌号 收购时间"
    输入平衡表-5月表，收油表-->输出收油表
    """
    def copy_balance_to_oil(df_generate_balance: pd.DataFrame, df_generate_oil: pd.DataFrame) -> pd.DataFrame:
        # 创建一个新的DataFrame df_generate_oil_copy 以避免直接修改原始df_generate_oil
        df_generate_oil_copy = df_generate_oil.copy()
        
        # 步骤1：从df_generate_balance复制流水号和交付时间到df_generate_oil的相应列
        # 首先需要基于车牌号和收集城市/区域进行关联
        merged_df = pd.merge(df_generate_oil_copy, df_generate_balance[['车牌号', '收集城市', '流水号', '交付时间']], 
                            left_on=['车牌号', '区域'], right_on=['车牌号', '收集城市'], how='left')
        
        # 更新df_generate_oil_copy的流水号和收购时间（即交付时间）
        df_generate_oil_copy['流水号'] = merged_df['流水号']
        df_generate_oil_copy['收购时间'] = merged_df['交付时间']
        
        return df_generate_oil_copy
    
    """
    平衡表-总表合同编号分配
    输入平衡表总表、收货确认书、当日生产转化系数、生成平衡表等的日期
    """
    def process_balance_sum_contract(df_generate_sum: pd.DataFrame, df_generate_check: pd.DataFrame,
                       df_generate_balance_last_month: pd.DataFrame, df_generate_balance_current_month: pd.DataFrame,
                       coeff_number: float, current_date: str):
        """
        处理两个DataFrame并生成一个新的DataFrame。
        
        :param df_generate_sum: 包含供应日期、产出重量、期末库存等信息的DataFrame
        :param df_generate_check: 包含重量列的DataFrame
        :param coeff_number: 浮点型数据，用于后续计算
        :param current_date: 字符串格式的当前日期
        :return: 处理后的DataFrame
        """
        # 步骤1：sum_product=df_generate_check的重量列进的和
        sum_product = df_generate_check['重量'].sum()
        
        # 步骤2：last_month=current_date上一个月；
        current_date = pd.to_datetime(current_date)
        last_month = (current_date - timedelta(days=current_date.day)).replace(day=1)
        last_month_ending_inventory = df_generate_sum[df_generate_sum['供应日期'].dt.to_period('M') == last_month.to_period('M')].iloc[-1]['期末库存']
        
        # 步骤3：当月的产出重量month_quantity = sum_product-last_month_ending_inventory 
        month_quantity = sum_product - last_month_ending_inventory
        
        # 步骤4：取出df_generate_sum表中供应日期的月份等于current_date月份并且产出重量不为空值或者null值的供应日期和产出重量值，并去重，
        # 循环去重后的供应日期和产出重量值，对产出重量值进行累加求和，当加到某一行的产出重量累加值<=month_quantity 并且下一行的产出重量累加值>month_quantity 时停止，
        # 记录当前行的供应日期stop_date和累加值sum_quantity；
        mask_current_month = (df_generate_sum['供应日期'].dt.to_period('M') == current_date.to_period('M')) & (df_generate_sum['产出重量'].notna())
        filtered_df = df_generate_sum[mask_current_month][['供应日期', '产出重量']].drop_duplicates()
        
        cumulative_sum = 0
        stop_date = None
        sum_quantity = 0
        for index, row in filtered_df.iterrows():
            cumulative_sum += row['产出重量']
            if cumulative_sum > month_quantity:
                stop_date = row['供应日期']
                sum_quantity = cumulative_sum - row['产出重量']
                break
            else:
                stop_date = filtered_df.iloc[-1]['供应日期']
                sum_quantity = cumulative_sum
        
        # 步骤5：求出剩余的原料remaining_materia= （month_quantity  -sum_quantity）/coeff_number,
        remaining_material = (month_quantity - sum_quantity) / coeff_number
        
        # 依次对df_generate_sum中供应日期大于stop_date的行的每车吨量累加求和，直到累加和>remaining_materia停止，记录对应的行索引stop_index
        mask_after_stop_date = (df_generate_sum['供应日期'] >= stop_date) #大于等于，因为stop_date上面是已经大于剩余量的日期
        cumulative_weight = 0
        stop_index = None
        for idx, row in df_generate_sum[mask_after_stop_date].iterrows():
            cumulative_weight += row['产出重量']
            if cumulative_weight > remaining_material:
                stop_index = idx
                break
        
        # 步骤6：填充df_generate_sum表中分配明细列，
        # 填充规则为1：供应日期的月份=current_date减1个月的合同分配明细列为空的分配明细列；
        # 2：供应日期的月份=current_date对应月份行索引<=stop_index的分配明细列。
        # 填充值为BWD-JC开头，加current_date年份的后2位数字，加current_date月份的第一天，
        # 例如current_date='2024-05-06'，则填充值为BWD-JC240501
        fill_value = f"BWD-JC{str(current_date.year)[-2:]}{current_date.month:02d}01"
        
        # 规则1
        mask_last_month = df_generate_sum['供应日期'].dt.to_period('M') == last_month.to_period('M')
        df_generate_sum.loc[mask_last_month & df_generate_sum['分配明细'].isna(), '分配明细'] = fill_value
        
        # 规则2
        mask_current_month = df_generate_sum['供应日期'].dt.to_period('M') == current_date.to_period('M')
        mask_until_stop_index = df_generate_sum.index <= stop_index
        df_generate_sum.loc[mask_current_month & mask_until_stop_index & df_generate_sum['分配明细'].isna(), '分配明细'] = fill_value
        # 步骤7：填充df_generate_balance_last_month表中合同分配明细列为空的分配明细列，值为BWD-JC开头，加current_date年份的后2位数字，加current_date月份的第一天；
        df_generate_balance_last_month.loc[df_generate_balance_last_month['分配明细'].isna(), '分配明细'] = fill_value
        
        # 步骤8：填充df_generate_balance_current_month分配明细列，规则为关联df_generate_sum表，根据日期和过磅单编号关联，
        # 填充值为BWD-JC开头，加current_date年份的后2位数字，加current_date月份的第一天
        # 假设关联字段是'供应日期' 和 '过磅单编号'
        df_generate_balance_current_month = df_generate_balance_current_month.merge(
            df_generate_sum[['供应日期', '分配明细']], left_on=['供应日期', '过磅单编号'], right_on=['供应日期', '过磅单编号'], how='left')
        df_generate_balance_current_month['分配明细_x'] = df_generate_balance_current_month['分配明细_y'].fillna(fill_value)
        df_generate_balance_current_month.drop(columns=['分配明细_y'], inplace=True)
        df_generate_balance_current_month.rename(columns={'分配明细_x': '分配明细'}, inplace=True)

        return df_generate_sum, df_generate_balance_last_month, df_generate_balance_current_month
    
    """
    复制流水号、交付时间和销售合同号
    输入收油表和平衡表_5月表"""
    def copy_balance_to_oil_dataframes(df_generate_oil: pd.DataFrame, df_generate_balance: pd.DataFrame) -> pd.DataFrame:
        """
        将df_generate_balance的信息合并到df_generate_oil中。
        
        :param df_generate_oil: 主表DataFrame，包含车牌号和累计收油数等信息
        :param df_generate_balance: 包含流水号、交付时间、销售合同号等信息的DataFrame
        :return: 合并后的df_generate_oil DataFrame
        """
        # 步骤1：将df_generate_oil和df_generate_balance关联，
        # 以df_generate_oil为主表，关联字段为车牌号和累计收油数，
        # 将df_generate_balance的流水号赋值给df_generate_oil的流水号，
        # 将df_generate_balance的交付时间赋值给df_generate_oil的收购时间，
        # 将df_generate_balance的销售合同号赋值给df_generate_oil的销售合同号
        
        # 假设关联是基于'车牌号'和'累计收油数'这两个字段
        merged_df = pd.merge(df_generate_oil, df_generate_balance[['车牌号', '累计收油数', '流水号', '交付时间', '销售合同号']],
                            on=['车牌号', '累计收油数'], how='left')
        
        # 更新df_generate_oil的相应列
        df_generate_oil['流水号'] = merged_df['流水号']
        df_generate_oil['收购时间'] = merged_df['交付时间']
        df_generate_oil['销售合同号'] = merged_df['销售合同号']
        
        return df_generate_oil